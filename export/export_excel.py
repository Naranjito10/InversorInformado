"""
Exporta listings a un .xlsx con codigo de colores por score_label.

Uso:
    python -m export.export_excel \
        --out viviendas.xlsx \
        --municipio "Barcelona" \
        --precio-min 100000 --precio-max 400000 \
        --score-min 50 --rentabilidad-min 5

Colores (segun prompt):
    alto         -> 🟢 verde   #22c55e
    medio        -> 🟡 amarillo #f59e0b
    incompleto   -> 🔴 rojo    #ef4444
    bajada_precio -> 🟣 morado  #a855f7 (solapado)
    normal       -> sin resaltar
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Permite usar este script como modulo y como script suelto
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scraper.db import query_listings  # noqa: E402


COLORS = {
    "alto":          "FF22C55E",  # verde
    "medio":         "FFF59E0B",  # amarillo
    "incompleto":    "FFEF4444",  # rojo
    "bajada_precio": "FFA855F7",  # morado (solo en columna bajada_precio)
}

FIELDS = [
    ("score",                "Score"),
    ("score_label",          "Label"),
    ("bajada_precio",        "Bajada"),
    ("fuente",               "Fuente"),
    ("titulo",               "Título"),
    ("precio_venta",         "Precio (€)"),
    ("precio_m2",            "€/m²"),
    ("precio_zona_m2",       "€/m² Zona"),
    ("descuento_zona_pct",   "% vs Zona"),
    ("metros_cuadrados",     "m²"),
    ("habitaciones",         "Hab"),
    ("banos",                "Baños"),
    ("planta",               "Planta"),
    ("ascensor",             "Ascensor"),
    ("terraza",              "Terraza"),
    ("garaje",               "Garaje"),
    ("estado",               "Estado"),
    ("certificado_energetico", "CEE"),
    ("alquiler_estimado",    "Alquiler (€/mes)"),
    ("rentabilidad_bruta",   "Rent. Bruta (%)"),
    ("rentabilidad_alquiler","Rent. Neta (€/año)"),
    ("ibi",                  "IBI (€/año)"),
    ("comunidad",            "Comunidad (€/año)"),
    ("derramas_pendientes",  "Derramas"),
    ("barrio",               "Barrio"),
    ("municipio",            "Municipio"),
    ("provincia",            "Provincia"),
    ("dias_en_mercado",      "Días mercado"),
    ("veces_visto",          "Veces visto"),
    ("primera_deteccion",    "Detectado"),
    ("ultima_actualizacion", "Actualizado"),
    ("activo",               "Activo"),
    ("url",                  "URL"),
]


def export_to_xlsx(filters: Optional[dict], out_path: str) -> str:
    rows = query_listings(filters or {}, limit=10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Viviendas"

    # Cabecera
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF1F2937")
    for col_idx, (_, label) in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Filas
    for r_idx, row in enumerate(rows, start=2):
        label = (row.get("score_label") or "normal").lower()
        bajada = bool(row.get("bajada_precio"))
        row_color = COLORS.get(label)

        for c_idx, (key, _) in enumerate(FIELDS, start=1):
            value = row.get(key)
            # Formateo de fechas
            if isinstance(value, str) and key in {"primera_deteccion", "ultima_actualizacion"}:
                try:
                    value = datetime.fromisoformat(value.replace("Z", "+00:00"))
                except Exception:
                    pass
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if row_color:
                cell.fill = PatternFill("solid", fgColor=row_color)
                # Texto en blanco para legibilidad sobre fondos saturados
                if label in {"alto", "incompleto"}:
                    cell.font = Font(color="FFFFFFFF")

            # Sobrescribir solo la columna `bajada_precio` con morado si aplica
            if bajada and key == "bajada_precio":
                cell.fill = PatternFill("solid", fgColor=COLORS["bajada_precio"])
                cell.font = Font(color="FFFFFFFF", bold=True)
                cell.value = "BAJADA ↓"

            if key == "url" and value:
                cell.hyperlink = value
                cell.font = Font(color="FF2563EB", underline="single")

    # Auto ancho de columnas (aprox)
    widths = {
        "score": 8, "score_label": 12, "bajada_precio": 10, "fuente": 12,
        "titulo": 50, "precio_venta": 14, "precio_m2": 10, "url": 40,
    }
    for c_idx, (key, _) in enumerate(FIELDS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(key, 14)

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)


def main() -> None:
    p = argparse.ArgumentParser(description="Exporta viviendas a Excel con colores por score.")
    p.add_argument("--out", default="export/viviendas.xlsx", help="Ruta del .xlsx de salida")
    p.add_argument("--municipio")
    p.add_argument("--barrio")
    p.add_argument("--fuente", choices=["idealista", "fotocasa", "habitaclia", "casaradar"])
    p.add_argument("--score-label", choices=["alto", "medio", "normal", "incompleto"])
    p.add_argument("--score-min", type=int)
    p.add_argument("--rentabilidad-min", type=float)
    p.add_argument("--precio-min", type=int)
    p.add_argument("--precio-max", type=int)
    p.add_argument("--fecha-desde")
    p.add_argument("--fecha-hasta")
    p.add_argument("--solo-activos", action="store_true", default=True)
    args = p.parse_args()

    filters = {
        "municipio": args.municipio,
        "barrio": args.barrio,
        "fuente": args.fuente,
        "score_label": args.score_label,
        "score_min": args.score_min,
        "rentabilidad_min": args.rentabilidad_min,
        "precio_min": args.precio_min,
        "precio_max": args.precio_max,
        "fecha_desde": args.fecha_desde,
        "fecha_hasta": args.fecha_hasta,
        "activo": True if args.solo_activos else None,
    }
    filters = {k: v for k, v in filters.items() if v is not None}

    out = export_to_xlsx(filters, args.out)
    print(f"Exportado: {out}")


if __name__ == "__main__":
    main()

"""
Exporta listings a un .xlsx con codigo de colores por score_label.

Colores (segun prompt):
    alto         -> verde   #22c55e
    medio        -> amarillo #f59e0b
    incompleto   -> rojo    #ef4444
    bajada_precio -> morado  #a855f7 (solapado)
    normal       -> sin resaltar
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scraper.infrastructure.db import query_listings


COLORS = {
    "alto":          "FF22C55E",  # verde
    "medio":         "FFF59E0B",  # amarillo
    "incompleto":    "FFEF4444",  # rojo
    "bajada_precio": "FFA855F7",  # morado (solo en columna bajada_precio)
}

FIELDS = [
    ("score",                "Score"),
    ("score_label",          "Label"),
    ("bajada_precio",        "Bajada"),
    ("fuente",               "Fuente"),
    ("titulo",               "Título"),
    ("precio_venta",         "Precio (€)"),
    ("precio_m2",            "€/m²"),
    ("precio_zona_m2",       "€/m² Zona"),
    ("descuento_zona_pct",   "% vs Zona"),
    ("metros_cuadrados",     "m²"),
    ("habitaciones",         "Hab"),
    ("banos",                "Baños"),
    ("planta",               "Planta"),
    ("ascensor",             "Ascensor"),
    ("terraza",              "Terraza"),
    ("garaje",               "Garaje"),
    ("condition",            "Estado"),
    ("certificado_energetico", "CEE"),
    ("alquiler_estimado",    "Alquiler (€/mes)"),
    ("rentabilidad_bruta",   "Rent. Bruta (%)"),
    ("rentabilidad_alquiler","Rent. Neta (€/año)"),
    ("ibi",                  "IBI (€/año)"),
    ("comunidad",            "Comunidad (€/año)"),
    ("derramas_pendientes",  "Derramas"),
    ("barrio",               "Barrio"),
    ("municipio",            "Municipio"),
    ("provincia",            "Provincia"),
    ("dias_en_mercado",      "Días mercado"),
    ("veces_visto",          "Veces visto"),
    ("primera_deteccion",    "Detectado"),
    ("ultima_actualizacion", "Actualizado"),
    ("status",               "Estado BD"),
    ("url",                  "URL"),
]


def generate_excel(out_path: Path, filters: Optional[dict] = None) -> str:
    """Genera el Excel con los listings filtrados y devuelve la ruta."""
    rows = query_listings(filters or {}, limit=10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Viviendas"

    # Cabecera
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF1F2937")
    for col_idx, (_, label) in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Filas
    for r_idx, row in enumerate(rows, start=2):
        label = (row.get("score_label") or "normal").lower()
        bajada = bool(row.get("bajada_precio"))
        row_color = COLORS.get(label)

        for c_idx, (key, _) in enumerate(FIELDS, start=1):
            value = row.get(key)
            # Formateo de fechas
            if isinstance(value, str) and key in {"primera_deteccion", "ultima_actualizacion"}:
                try:
                    value = datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
                except Exception:
                    pass
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if row_color:
                cell.fill = PatternFill("solid", fgColor=row_color)
                if label in {"alto", "incompleto"}:
                    cell.font = Font(color="FFFFFFFF")

            # Sobrescribir solo la columna `bajada_precio` con morado si aplica
            if bajada and key == "bajada_precio":
                cell.fill = PatternFill("solid", fgColor=COLORS["bajada_precio"])
                cell.font = Font(color="FFFFFFFF", bold=True)
                cell.value = "BAJADA ↓"

            if key == "url" and value:
                cell.hyperlink = value
                cell.font = Font(color="FF2563EB", underline="single")

    # Auto ancho de columnas (aprox)
    widths = {
        "score": 8, "score_label": 12, "bajada_precio": 10, "fuente": 12,
        "titulo": 50, "precio_venta": 14, "precio_m2": 10, "url": 40,
    }
    for c_idx, (key, _) in enumerate(FIELDS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(key, 14)

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return str(out_path)